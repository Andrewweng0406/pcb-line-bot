from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from datetime import datetime
import os


quote_counter = 1


def export_quote_excel(parsed, result):

    global quote_counter

    wb = Workbook()
    ws = wb.active

    ws.title = "PCB Quote"

    quote_no = f"PCB-{datetime.now().strftime('%Y%m%d')}-{quote_counter:03d}"

    quote_counter += 1

    # Title
    ws.merge_cells("A1:D1")

    ws["A1"] = "PCB 正式報價單"

    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Company
    ws["A3"] = "Company"
    ws["B3"] = "Your PCB Company"

    ws["A4"] = "Quote No"
    ws["B4"] = quote_no

    ws["A5"] = "Date"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Basic Spec
    ws["A7"] = "Layer"
    ws["B7"] = parsed.get("layer")

    ws["A8"] = "Material"
    ws["B8"] = parsed.get("material")

    ws["A9"] = "Size"

    if parsed.get("length_mm") and parsed.get("width_mm"):
        ws["B9"] = f'{parsed.get("length_mm")} x {parsed.get("width_mm")} mm'
    else:
        ws["B9"] = f'{parsed.get("area_inch")} sq.inch'

    ws["A10"] = "Area"
    ws["B10"] = result.get("area_inch")

    ws["A11"] = "Qty"
    ws["B11"] = parsed.get("qty")

    # Process
    ws["A13"] = "ENIG"
    ws["B13"] = str(parsed.get("enig"))

    ws["A14"] = "ENIG Thickness"
    ws["B14"] = str(parsed.get("enig_thickness_uinch"))

    ws["A15"] = "VIP"
    ws["B15"] = str(parsed.get("vip"))

    ws["A16"] = "Impedance"
    ws["B16"] = str(parsed.get("impedance"))

    ws["A17"] = "Back Drill"
    ws["B17"] = str(parsed.get("back_drill"))

    ws["A18"] = "BVH"
    ws["B18"] = str(parsed.get("bvh"))

    # Price Details
    ws["A20"] = "Setup Fee (Engineering)"
    ws["B20"] = result.get("setup_fee", result.get("engineering_fee"))

    ws["A21"] = "Board Charge (Material)"
    ws["B21"] = result.get("board_charge_total", result.get("material_cost"))

    ws["A22"] = "Extra Fee (Process)"
    ws["B22"] = result.get("extra_fee", result.get("process_cost"))

    ws["A23"] = "Subtotal"
    ws["B23"] = result.get("subtotal")

    ws["A24"] = "Discount"
    ws["B24"] = result.get("discount")

    ws["A25"] = "Delivery Multiplier"
    ws["B25"] = result.get("delivery_multiplier", 1.0)

    ws["A26"] = "TOTAL"

    ws["B26"] = result.get("total")

    ws["A26"].font = Font(size=14, bold=True)
    ws["B26"].font = Font(size=14, bold=True)

    ws["A26"].fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    )

    ws["B26"].fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    )

    # Width
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30

    filename = f'{quote_no}.xlsx'

    # Ensure the exports directory exists
    exports_dir = "exports"
    if not os.path.exists(exports_dir):
        os.makedirs(exports_dir)

    # Save to the exports directory
    filepath = os.path.join(exports_dir, filename)
    wb.save(filepath)

    return filename
