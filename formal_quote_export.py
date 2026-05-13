from openpyxl import load_workbook
from datetime import datetime
import os


def export_formal_quote(parsed, result):
    template_path = "templates/quote_template.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    # 基本資料
    ws["C5"] = "ProbeLeader CO., LTD."
    ws["C6"] = "AI Quote System"
    ws["C7"] = datetime.now().strftime("%Y/%m/%d")

    # 規格
    ws["F10"] = parsed.get("layer")

    if parsed.get("length_mm") and parsed.get("width_mm"):
        ws["F11"] = parsed.get("length_mm")
        ws["H11"] = parsed.get("width_mm")

    else:
        ws["F11"] = result.get("area_inch")
        ws["H11"] = "sq.inch"

    ws["F14"] = parsed.get("material")
    ws["F15"] = parsed.get("thickness")
    ws["F16"] = parsed.get("copper_weight")

    if parsed.get("enig"):
        ws["F17"] = f'Gold {parsed.get("enig_thickness_uinch")}u"'
    else:
        ws["F17"] = parsed.get("surface_finish")

    ws["F18"] = "YES" if parsed.get("gold_finger") else "NO"

    comment = []

    if parsed.get("vip"):
        comment.append("VIP")
    if parsed.get("back_drill"):
        comment.append("Back Drill")
    if parsed.get("bvh"):
        comment.append("BVH")
    if parsed.get("impedance"):
        comment.append("Impedance: 50ohm")

    ws["F19"] = " / ".join(comment)

    # 價格
    ws["J14"] = parsed.get("qty")
    ws["K14"] = "pcs/batch"
    ws["L14"] = f'NT${result.get("unit_price")} / pc'

    # 交期
    if result.get("delivery_days"):
        ws["N14"] = result.get("delivery_days")
        ws["O14"] = "working days"
    else:
        ws["N14"] = ""
        ws["O14"] = "working days"

    filename = f'formal_quote_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    output_path = os.path.join("exports", filename)

    wb.save(output_path)

    return output_path